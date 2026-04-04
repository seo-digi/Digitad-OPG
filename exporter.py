"""Export before/after SEO optimization data to .xlsx."""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


HEADERS = [
    "URL",
    "Mot-clé cible",
    "Volume mensuel",
    "Position actuelle",
    "Ancien Title",
    "Nouveau Title",
    "Ancien H1",
    "Nouveau H1",
    "Ancienne Meta Desc",
    "Nouvelle Meta Desc",
]

GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

OLD_COLS = {5, 7, 9}
NEW_COLS = {6, 8, 10}


def export_xlsx(data, output_path):
    """Export optimization data to formatted .xlsx."""
    wb = Workbook()
    first_sheet = True

    for lang in ["fr", "en"]:
        if lang not in data or not data[lang]:
            continue

        if first_sheet:
            ws = wb.active
            ws.title = lang.upper()
            first_sheet = False
        else:
            ws = wb.create_sheet(lang.upper())

        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for row_idx, page in enumerate(data[lang], 2):
            values = [
                page.get("url", ""),
                page.get("keyword", ""),
                page.get("volume", ""),
                page.get("position", ""),
                page.get("old_title", ""),
                page.get("new_title", ""),
                page.get("old_h1", ""),
                page.get("new_h1", ""),
                page.get("old_meta_desc", ""),
                page.get("new_meta_desc", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if col in OLD_COLS:
                    cell.fill = GREY_FILL
                elif col in NEW_COLS:
                    cell.fill = GREEN_FILL

        for col in range(1, len(HEADERS) + 1):
            max_len = len(HEADERS[col - 1])
            for row in range(2, len(data[lang]) + 2):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 60))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 2

    wb.save(output_path)
