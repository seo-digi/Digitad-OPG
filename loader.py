"""Load and normalize GSC CSV and keyword study Excel data."""
import re
import pandas as pd


def _parse_volume(val):
    """Parse volume value, handling space/comma thousand separators (e.g., '4 400')."""
    if not val or val == "-":
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    cleaned = str(val).strip().replace(" ", "").replace("\u00a0", "").replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        return 0
from openpyxl import load_workbook
from config import (
    KEYWORD_COLUMN_LABELS, VOLUME_COLUMN_LABELS, POSITION_COLUMN_LABELS,
    PRIORITY_COLUMN_LABELS, URL_COLUMN_LABELS, INTENT_COLUMN_LABELS,
    DEFAULT_HEADER_ROW, DEFAULT_DATA_START_ROW,
)


def _matches_labels(val, labels):
    """Return True if val matches any label in labels.

    For single-word labels (no spaces), require a whole-word match to avoid
    false positives (e.g. "position" matching "url positionnée").
    For multi-word labels, a substring match is sufficient.
    """
    for label in labels:
        if " " in label:
            if label in val:
                return True
        else:
            # Whole-word match using word boundaries
            if re.search(r"\b" + re.escape(label) + r"\b", val):
                return True
    return False


def load_gsc_csv(csv_path):
    """Load SEO Gets CSV, normalize column names and URLs."""
    df = pd.read_csv(csv_path)
    # Normalize column names to lowercase with underscores
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    # Ensure numeric types
    for col in ["clicks", "impressions", "ctr", "avg_position"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    # Normalize URLs: strip trailing slashes, lowercase
    df["page"] = df["page"].str.rstrip("/").str.lower()
    return df


def detect_columns(xlsx_path, sheet_name):
    """Auto-detect keyword, volume, position, priority, URL, and intent columns.

    Scans the header row for known labels (FR and EN variants).
    Returns a dict with column indices and row numbers.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[sheet_name]

    result = {
        "keyword_col": None,
        "volume_col": None,
        "position_col": None,
        "priority_col": None,
        "url_col": None,
        "intent_col": None,
        "header_row": DEFAULT_HEADER_ROW,
        "data_start_row": DEFAULT_DATA_START_ROW,
    }

    for row_idx in range(max(1, DEFAULT_HEADER_ROW - 2), DEFAULT_HEADER_ROW + 3):
        for cell in ws[row_idx]:
            if cell.value is None:
                continue
            val = str(cell.value).strip().lower().replace("\n", " ")
            # Check specific labels before the broad keyword catch-all to avoid
            # false positives (e.g. "Difficulté de mot-clé (KD)" contains "mot-clé",
            # or "position" matching "url positionnée").
            # Use first-match-wins (only assign if slot is still None).
            if _matches_labels(val, VOLUME_COLUMN_LABELS):
                if result["volume_col"] is None:
                    result["volume_col"] = cell.column
            elif _matches_labels(val, POSITION_COLUMN_LABELS):
                if result["position_col"] is None:
                    result["position_col"] = cell.column
            elif _matches_labels(val, PRIORITY_COLUMN_LABELS):
                if result["priority_col"] is None:
                    result["priority_col"] = cell.column
            elif _matches_labels(val, URL_COLUMN_LABELS) and "positionnée" in val:
                if result["url_col"] is None:
                    result["url_col"] = cell.column
            elif _matches_labels(val, INTENT_COLUMN_LABELS):
                if result["intent_col"] is None:
                    result["intent_col"] = cell.column
            elif _matches_labels(val, KEYWORD_COLUMN_LABELS):
                if result["keyword_col"] is None:
                    result["keyword_col"] = cell.column
                    result["header_row"] = row_idx
                    result["data_start_row"] = row_idx + 1

    wb.close()
    return result


def load_keyword_study(xlsx_path, sheet_name, col_map):
    """Load keyword study data using detected column mapping."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[sheet_name]

    keywords = []
    for row in ws.iter_rows(min_row=col_map["data_start_row"], values_only=False):
        kw_cell = row[col_map["keyword_col"] - 1]
        if kw_cell.value is None:
            continue

        vol_val = row[col_map["volume_col"] - 1].value if col_map["volume_col"] else None
        pos_val = row[col_map["position_col"] - 1].value if col_map["position_col"] else None
        pri_val = row[col_map["priority_col"] - 1].value if col_map.get("priority_col") else None
        url_val = row[col_map["url_col"] - 1].value if col_map.get("url_col") else None
        intent_val = row[col_map["intent_col"] - 1].value if col_map.get("intent_col") else None

        keywords.append({
            "keyword": str(kw_cell.value).strip(),
            "volume": _parse_volume(vol_val),
            "position": str(pos_val).strip() if pos_val else "-",
            "priority": str(pri_val).strip() if pri_val else "",
            "url": str(url_val).strip() if url_val and url_val != "-" else "",
            "intent": str(intent_val).strip() if intent_val else "",
        })

    wb.close()
    return pd.DataFrame(keywords)


def split_by_language(df, en_pattern="/en/"):
    """Split a GSC dataframe into FR and EN based on URL pattern."""
    en_mask = df["page"].str.contains(en_pattern, case=False, na=False)
    return df[~en_mask].copy(), df[en_mask].copy()
