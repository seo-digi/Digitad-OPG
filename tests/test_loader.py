"""Tests for loader module."""
import pandas as pd
import pytest
import os
import tempfile
from openpyxl import Workbook

from loader import load_gsc_csv, load_keyword_study, detect_columns, split_by_language


@pytest.fixture
def sample_csv(tmp_path):
    """Create a sample SEO Gets CSV."""
    csv_path = tmp_path / "test.csv"
    csv_path.write_text(
        '"page","query","Clicks","Impressions","CTR","Avg Position"\n'
        '"https://example.com/service/derm/","dermatologue",66,6944,0.95,5.05\n'
        '"https://example.com/en/service/derm/","dermatologist",30,2000,1.5,4.0\n'
        '"https://example.com/","clinique",45,3450,1.3,8.3\n'
    )
    return str(csv_path)


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a sample keyword study Excel file."""
    xlsx_path = tmp_path / "keywords.xlsx"
    wb = Workbook()

    # FR sheet
    ws_fr = wb.active
    ws_fr.title = "ÉTUDE DE MOTS-CLÉS - FR"
    ws_fr.cell(row=2, column=2, value="ÉTUDE DE MOTS CLÉS")
    ws_fr.cell(row=4, column=2, value="Mot-clé")
    ws_fr.cell(row=4, column=3, value="Volume de recherche mensuel")
    ws_fr.cell(row=4, column=4, value="Difficulté de mot-clé (KD)")
    ws_fr.cell(row=4, column=5, value="Intention\nde recherche")
    ws_fr.cell(row=4, column=6, value="Position actuelle")
    ws_fr.cell(row=4, column=7, value="URL positionnée")
    ws_fr.cell(row=4, column=8, value="Priorité stratégique")
    ws_fr.cell(row=5, column=2, value="dermatologue")
    ws_fr.cell(row=5, column=3, value=6600)
    ws_fr.cell(row=5, column=6, value="12")
    ws_fr.cell(row=5, column=7, value="https://example.com/derm")
    ws_fr.cell(row=5, column=8, value="high")
    ws_fr.cell(row=6, column=2, value="clinique privée")
    ws_fr.cell(row=6, column=3, value=3200)
    ws_fr.cell(row=6, column=6, value="-")

    # EN sheet
    ws_en = wb.create_sheet("ÉTUDE DE MOTS-CLÉS - EN")
    ws_en.cell(row=4, column=2, value="Mot-clé")
    ws_en.cell(row=4, column=3, value="Volume de recherche mensuel")
    ws_en.cell(row=4, column=6, value="Position actuelle")
    ws_en.cell(row=5, column=2, value="dermatologist montreal")
    ws_en.cell(row=5, column=3, value=2900)
    ws_en.cell(row=5, column=6, value="8")

    wb.save(xlsx_path)
    return str(xlsx_path)


class TestLoadGscCsv:
    def test_loads_csv_and_normalizes_urls(self, sample_csv):
        df = load_gsc_csv(sample_csv)
        assert len(df) == 3
        assert df.iloc[0]["page"] == "https://example.com/service/derm"
        assert df.iloc[0]["impressions"] == 6944
        assert df.iloc[0]["ctr"] == 0.95

    def test_handles_ctr_zero(self, tmp_path):
        csv_path = tmp_path / "zero_ctr.csv"
        csv_path.write_text(
            '"page","query","Clicks","Impressions","CTR","Avg Position"\n'
            '"https://example.com/","test",0,100,0,8.0\n'
        )
        df = load_gsc_csv(str(csv_path))
        assert df.iloc[0]["ctr"] == 0.0


class TestDetectColumns:
    def test_detects_fr_columns(self, sample_xlsx):
        result = detect_columns(sample_xlsx, "ÉTUDE DE MOTS-CLÉS - FR")
        assert result["keyword_col"] == 2
        assert result["volume_col"] == 3
        assert result["position_col"] == 6
        assert result["priority_col"] == 8
        assert result["url_col"] == 7
        assert result["intent_col"] == 5
        assert result["header_row"] == 4
        assert result["data_start_row"] == 5


class TestLoadKeywordStudy:
    def test_loads_keywords_with_all_fields(self, sample_xlsx):
        col_map = {
            "keyword_col": 2, "volume_col": 3,
            "position_col": 6, "priority_col": 8,
            "url_col": 7, "intent_col": 5,
            "header_row": 4, "data_start_row": 5
        }
        df = load_keyword_study(sample_xlsx, "ÉTUDE DE MOTS-CLÉS - FR", col_map)
        assert len(df) == 2
        assert df.iloc[0]["keyword"] == "dermatologue"
        assert df.iloc[0]["volume"] == 6600
        assert df.iloc[0]["position"] == "12"
        assert df.iloc[0]["priority"] == "high"
        assert df.iloc[0]["url"] == "https://example.com/derm"


class TestSplitByLanguage:
    def test_splits_urls_by_en_pattern(self, sample_csv):
        df = load_gsc_csv(sample_csv)
        fr_df, en_df = split_by_language(df, en_pattern="/en/")
        assert len(fr_df) == 2
        assert len(en_df) == 1
        assert "/en/" in en_df.iloc[0]["page"]
