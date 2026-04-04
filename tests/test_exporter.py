"""Tests for exporter module."""
import os
import pytest
from openpyxl import load_workbook

from exporter import export_xlsx


@pytest.fixture
def sample_data():
    return {
        "fr": [
            {
                "url": "https://example.com/derm",
                "keyword": "dermatologue",
                "volume": 6600,
                "position": "12",
                "old_title": "Dermatologie - Brand",
                "new_title": "Dermatologue : Soins experts - Brand",
                "old_h1": "Dermatologie",
                "new_h1": "Dermatologue a Blainville",
                "old_meta_desc": "Old description.",
                "new_meta_desc": "New description here.",
            },
        ],
        "en": [
            {
                "url": "https://example.com/en/derm",
                "keyword": "dermatologist",
                "volume": 2900,
                "position": "8",
                "old_title": "Dermatology - Brand",
                "new_title": "Dermatologist Montreal: Expert Care - Brand",
                "old_h1": "Dermatology",
                "new_h1": "Dermatologist in Montreal",
                "old_meta_desc": "Old EN desc.",
                "new_meta_desc": "New EN description here.",
            },
        ],
    }


class TestExportXlsx:
    def test_creates_xlsx_with_two_sheets(self, sample_data, tmp_path):
        output_path = str(tmp_path / "test_output.xlsx")
        export_xlsx(sample_data, output_path)

        assert os.path.exists(output_path)
        wb = load_workbook(output_path)
        assert "FR" in wb.sheetnames
        assert "EN" in wb.sheetnames

    def test_correct_column_headers(self, sample_data, tmp_path):
        output_path = str(tmp_path / "test_output.xlsx")
        export_xlsx(sample_data, output_path)

        wb = load_workbook(output_path)
        ws = wb["FR"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 11)]
        assert "URL" in headers
        assert "Mot-clé cible" in headers
        assert "Nouveau Title" in headers

    def test_data_rows_populated(self, sample_data, tmp_path):
        output_path = str(tmp_path / "test_output.xlsx")
        export_xlsx(sample_data, output_path)

        wb = load_workbook(output_path)
        ws = wb["FR"]
        assert ws.cell(row=2, column=1).value == "https://example.com/derm"
        assert ws.cell(row=2, column=2).value == "dermatologue"
