"""Digitad OPG — Optimization Plan Generator."""
import argparse
import json
import os
import sys
import datetime

# Load .env file if present
_env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _key, _val = _line.split("=", 1)
                os.environ.setdefault(_key.strip(), _val.strip())
from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt, Confirm, IntPrompt

from config import build_client_config, SIMILARITY_THRESHOLD, MAX_REGEN_ATTEMPTS, SCRAPER_RATE_LIMIT
from server import ValidationServer
from loader import load_gsc_csv, detect_columns, load_keyword_study, split_by_language
from mapper import (
    aggregate_gsc_by_url, get_top_queries_per_url,
    compute_opportunity_scores, map_keywords_to_urls,
    reconcile_unmapped_urls,
)
from scraper import scrape_tags
from rewriter import validate_mapping, rewrite_tags, validate_meta_desc_length
from exporter import export_xlsx

console = Console()

# Page type inference from URL patterns
PAGE_TYPE_PATTERNS = {
    "blog": ["/blog/", "/article/", "/actualite/", "/news/"],
    "service": ["/service/", "/services/"],
    "product": ["/product/", "/produit/"],
    "landing": ["/landing/", "/lp/"],
    "contact": ["/contact"],
    "career": ["/carriere", "/career"],
}


def infer_page_type(url, intent=""):
    """Infer page type from URL patterns and search intent."""
    url_lower = url.lower()
    for ptype, patterns in PAGE_TYPE_PATTERNS.items():
        if any(p in url_lower for p in patterns):
            return ptype
    if intent:
        intent_lower = intent.lower()
        if "informational" in intent_lower:
            return "blog"
        if "commercial" in intent_lower:
            return "service"
    path = url_lower.rstrip("/").split("/")
    if len(path) <= 3:
        return "homepage"
    return "service"


def get_checkpoint_path(client_name, step):
    return f".checkpoints/{client_name}_step{step}.json"


def save_checkpoint(client_name, step, data):
    os.makedirs(".checkpoints", exist_ok=True)
    path = get_checkpoint_path(client_name, step)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    console.print(f"[dim]Checkpoint saved: {path}[/dim]")


def load_checkpoint(client_name, step):
    path = get_checkpoint_path(client_name, step)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return None


def display_column_detection(col_map, sheet_name):
    """Display detected columns for consultant validation."""
    from openpyxl.utils import get_column_letter
    console.print(f"\n[bold]Detected columns in sheet \"{sheet_name}\":[/bold]")
    for key, label in [("keyword_col", "Keywords"), ("volume_col", "Volume"),
                       ("position_col", "Position"), ("priority_col", "Priority"),
                       ("url_col", "URL"), ("intent_col", "Intent")]:
        if col_map.get(key):
            console.print(f"  {label}:  Column {get_column_letter(col_map[key])} (col {col_map[key]})")
    console.print(f"  Start row: {col_map['data_start_row']}")


def display_mapping_table(mapping, opportunity_scores):
    """Display keyword mapping for consultant review."""
    table = Table(title="Keyword Mapping")
    table.add_column("#", style="dim")
    table.add_column("URL", max_width=45)
    table.add_column("Keyword", style="bold")
    table.add_column("Similarity", justify="right")
    table.add_column("Volume", justify="right")
    table.add_column("Opportunity", justify="right")
    table.add_column("Confidence")
    table.add_column("Top GSC Queries", max_width=40)

    sorted_urls = sorted(mapping.keys(),
                         key=lambda u: opportunity_scores.get(u, 0),
                         reverse=True)

    for i, url in enumerate(sorted_urls, 1):
        m = mapping[url]
        short_url = url.replace("https://", "").replace("http://", "")
        top_q = ", ".join(q["query"] for q in m.get("top_queries", [])[:3])
        conf = m.get("confidence", "OK")
        conf_style = "green" if conf == "OK" else ("yellow" if conf == "STUDY" else "red")

        table.add_row(
            str(i), short_url, m["keyword"],
            f"{m['similarity']:.2f}",
            str(m.get("volume", "")),
            f"{opportunity_scores.get(url, 0):.0f}",
            f"[{conf_style}]{conf}[/{conf_style}]",
            top_q,
        )

    console.print(table)


def display_rewrite_preview(page_data):
    """Display before/after preview for a single page."""
    console.print(f"\n[bold]URL:[/bold] {page_data['url']}")
    console.print(f"  [dim]Title:[/dim]  {page_data['old_title']}")
    console.print(f"  [green]Title:[/green]  {page_data['new_title']}")
    console.print(f"  [dim]H1:[/dim]     {page_data['old_h1']}")
    console.print(f"  [green]H1:[/green]     {page_data['new_h1']}")
    console.print(f"  [dim]Meta:[/dim]   {page_data['old_meta_desc']}")
    console.print(f"  [green]Meta:[/green]   {page_data['new_meta_desc']} ({len(page_data['new_meta_desc'])} chars)")


def interactive_mapping_review(mapping, opportunity_scores):
    """Let consultant review and modify keyword mappings."""
    display_mapping_table(mapping, opportunity_scores)

    if Confirm.ask("\nApprove all mappings?"):
        return mapping

    console.print("[yellow]Review each mapping. Enter new keyword, SKIP to remove, or Enter to accept.[/yellow]")
    sorted_urls = sorted(list(mapping.keys()),
                         key=lambda u: opportunity_scores.get(u, 0),
                         reverse=True)

    to_remove = []
    for url in sorted_urls:
        m = mapping[url]
        short_url = url.replace("https://", "").replace("http://", "")
        new_kw = Prompt.ask(
            f"  {short_url} -> [bold]{m['keyword']}[/bold]",
            default=m["keyword"],
        )
        if new_kw == "SKIP":
            to_remove.append(url)
        elif new_kw != m["keyword"]:
            mapping[url]["keyword"] = new_kw

    for url in to_remove:
        del mapping[url]

    return mapping


def regenerate_single_page(page_payload, brand_name, lang, config, guidance=""):
    """Regenerate rewrites for a single page with optional guidance."""
    if guidance:
        page_payload["guidance"] = guidance
    results = rewrite_tags([page_payload], brand_name, lang, model=config["model_rewriting"])
    return results[0] if results else None


def main():
    parser = argparse.ArgumentParser(description="Digitad SEO Optimization Plan Generator")
    parser.add_argument("--resume", action="store_true", help="Resume from last checkpoint")
    args = parser.parse_args()

    console.print("[bold blue]═══ Digitad OPG — Optimization Plan Generator ═══[/bold blue]\n")

    # --- Check API key ---
    if not os.environ.get("ANTHROPIC_API_KEY"):
        console.print("[bold red]ERREUR: Cle API Anthropic manquante.[/bold red]")
        console.print("Copiez le fichier .env.example en .env et ajoutez votre cle:")
        console.print("  cp .env.example .env")
        console.print("  # Editez .env et ajoutez votre ANTHROPIC_API_KEY")
        sys.exit(1)

    # --- Client config ---
    brand_name = Prompt.ask("Brand name", default="Medicym")
    languages_input = Prompt.ask("Languages (fr, en, or both)", default="both")
    languages = ["fr", "en"] if languages_input == "both" else [languages_input]
    config = build_client_config(brand_name, languages)
    client_name = brand_name.lower().replace(" ", "_")

    # Check for resume
    resume = args.resume
    if not resume and os.path.exists(get_checkpoint_path(client_name, 2)):
        resume = Confirm.ask("Checkpoint found. Resume from last step?")

    # --- Step 1: Load data ---
    console.print("\n[bold]Step 1: Loading data...[/bold]")
    gsc_csv_path = Prompt.ask("Path to SEO Gets CSV")
    gsc_df = load_gsc_csv(gsc_csv_path)
    console.print(f"  Loaded {len(gsc_df)} GSC rows, {gsc_df['page'].nunique()} unique pages")

    xlsx_path = Prompt.ask("Path to keyword study Excel")

    all_keywords = {}
    for lang in languages:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True)
        sheet_name = None
        for name in wb.sheetnames:
            if lang.upper() in name.upper() and ("MOT" in name.upper() or "KEY" in name.upper()):
                sheet_name = name
                break
        wb.close()

        if not sheet_name:
            console.print(f"[yellow]No keyword sheet found for {lang.upper()}, skipping[/yellow]")
            continue

        col_map = detect_columns(xlsx_path, sheet_name)
        display_column_detection(col_map, sheet_name)

        if not Confirm.ask("Is this correct?"):
            console.print("Enter column numbers manually:")
            col_map["keyword_col"] = IntPrompt.ask("Keyword column number")
            col_map["volume_col"] = IntPrompt.ask("Volume column number")
            col_map["position_col"] = IntPrompt.ask("Position column number (0 to skip)")
            col_map["data_start_row"] = IntPrompt.ask("Data start row")
            if col_map["position_col"] == 0:
                col_map["position_col"] = None

        kw_df = load_keyword_study(xlsx_path, sheet_name, col_map)
        all_keywords[lang] = kw_df
        console.print(f"  Loaded {len(kw_df)} keywords for {lang.upper()}")

    fr_gsc, en_gsc = split_by_language(gsc_df, en_pattern=config["en_url_pattern"])
    num_pages = IntPrompt.ask("Number of pages to optimize (0 = all)", default=0)

    # --- Step 2: Keyword mapping ---
    if resume and (checkpoint := load_checkpoint(client_name, 2)):
        mapping = checkpoint
        console.print("[green]Loaded mapping from checkpoint[/green]")
    else:
        console.print("\n[bold]Step 2: Mapping keywords to URLs (embedding model)...[/bold]")

        from sentence_transformers import SentenceTransformer
        from config import EMBEDDING_MODEL
        console.print(f"  Loading model {EMBEDDING_MODEL}...")
        model = SentenceTransformer(EMBEDDING_MODEL)

        mapping = {}
        for lang in languages:
            lang_gsc = fr_gsc if lang == "fr" else en_gsc
            if lang not in all_keywords or lang_gsc.empty:
                continue
            lang_mapping = map_keywords_to_urls(lang_gsc, all_keywords[lang], model=model)
            for url, m in lang_mapping.items():
                m["lang"] = lang
            mapping.update(lang_mapping)
            console.print(f"  Mapped {len(lang_mapping)} {lang.upper()} pages")

            unmapped = reconcile_unmapped_urls(all_keywords[lang], set(mapping.keys()))
            if unmapped:
                console.print(f"  [yellow]{len(unmapped)} keyword study URLs have no GSC data:[/yellow]")
                for u in unmapped:
                    u["lang"] = lang
                    mapping[u["url"]] = u
                    console.print(f"    {u['url']} -> {u['keyword']} (from study)")

        save_checkpoint(client_name, 2, mapping)

    opp_scores = compute_opportunity_scores(gsc_df)

    if num_pages > 0:
        sorted_urls = sorted(mapping.keys(), key=lambda u: opp_scores.get(u, 0), reverse=True)
        mapping = {u: mapping[u] for u in sorted_urls[:num_pages]}

    # --- Step 3: AI validation ---
    console.print("\n[bold]Step 3: AI validation of keyword mapping...[/bold]")
    if Confirm.ask("Run AI validation? (uses tokens)"):
        mapping_list = [
            {"url": url, "keyword": m["keyword"], "top_queries": m.get("top_queries", [])}
            for url, m in mapping.items()
        ]
        validated = validate_mapping(mapping_list, model=config["model_validation"])
        for v in validated:
            if v["url"] in mapping:
                mapping[v["url"]]["ai_status"] = v["status"]
                if v["suggestion"]:
                    mapping[v["url"]]["ai_suggestion"] = v["suggestion"]
                    mapping[v["url"]]["ai_reason"] = v.get("reason", "")
        console.print("[green]Validation complete[/green]")
    else:
        console.print("[yellow]Skipped AI validation[/yellow]")

    # --- Step 4: Consultant validation (browser) ---
    console.print("\n[bold]Step 4: Validation du mapping dans le navigateur[/bold]")
    server = ValidationServer(mapping, opp_scores, config)
    url = server.start()
    console.print(f"[bold]Page de validation ouverte : {url}[/bold]")
    console.print("[dim]En attente de la validation dans le navigateur...[/dim]")
    mapping = server.wait_for_validation()
    console.print(f"\n[green]✓ {len(mapping)} pages confirmees[/green]")

    # --- Step 5: Scrape current tags ---
    if resume and (checkpoint := load_checkpoint(client_name, 5)):
        scraped = checkpoint
        console.print("[green]Loaded scraped data from checkpoint[/green]")
    else:
        console.print(f"\n[bold]Step 5: Scraping current tags from {len(mapping)} pages...[/bold]")
        server.complete_step("validation")
        urls_to_scrape = list(mapping.keys())
        server.update_progress("scraping", 0, len(urls_to_scrape))
        scraped = {}
        for i, scrape_url in enumerate(urls_to_scrape):
            result = scrape_tags([scrape_url], rate_limit=0 if i == 0 else SCRAPER_RATE_LIMIT)
            scraped.update(result)
            server.update_progress("scraping", i + 1, len(urls_to_scrape))
        save_checkpoint(client_name, 5, scraped)

        errors = [url for url, s in scraped.items() if "error" in s]
        if errors:
            console.print(f"[yellow]Warning: {len(errors)} pages had scraping errors:[/yellow]")
            for url in errors:
                console.print(f"  {url}: {scraped[url]['error']}")

    console.print("[green]Scraping complete[/green]")

    # --- Step 6: Rewrite tags ---
    console.print("\n[bold]Step 6: Rewriting SEO tags via Claude API...[/bold]")
    server.complete_step("scraping")

    # Count total pages across all languages for progress
    all_rewrite_pages = []
    for lang in languages:
        for url, m in mapping.items():
            if m.get("lang") == lang:
                all_rewrite_pages.append(url)
    total_rewrite = len(all_rewrite_pages)
    rewrite_done = 0

    rewrite_results = {}
    for lang in languages:
        pages = []
        for url, m in mapping.items():
            if m.get("lang") != lang:
                continue
            tags = scraped.get(url, {})
            intent = m.get("intent", "")
            pages.append({
                "url": url,
                "keyword": m["keyword"],
                "page_type": infer_page_type(url, intent),
                "current_title": tags.get("title", ""),
                "current_h1": tags.get("h1", ""),
                "current_meta_desc": tags.get("meta_description", ""),
            })

        if not pages:
            continue

        console.print(f"  Rewriting {len(pages)} {lang.upper()} pages...")
        server.update_progress("rewriting", rewrite_done, total_rewrite)

        def on_batch(completed, _total, _offset=rewrite_done):
            server.update_progress("rewriting", _offset + completed, total_rewrite)

        results = rewrite_tags(pages, brand_name, lang, model=config["model_rewriting"],
                               on_batch_complete=on_batch)
        rewrite_done += len(results)
        server.update_progress("rewriting", rewrite_done, total_rewrite)

        for i, r in enumerate(results):
            if not validate_meta_desc_length(r["new_meta_desc"]):
                console.print(f"  [yellow]Meta desc for {r['url']} is {len(r['new_meta_desc'])} chars, regenerating...[/yellow]")
                for attempt in range(MAX_REGEN_ATTEMPTS):
                    new_r = regenerate_single_page(pages[i], brand_name, lang, config,
                                                   guidance="Meta description must be 145-155 characters exactly.")
                    if new_r and validate_meta_desc_length(new_r["new_meta_desc"]):
                        results[i] = new_r
                        break

        rewrite_results[lang] = results

    save_checkpoint(client_name, 6, rewrite_results)

    # --- Step 7: Consultant review ---
    console.print("\n[bold]Step 7: Review rewrites[/bold]")
    export_data = {"fr": [], "en": []}

    for lang, results in rewrite_results.items():
        pages_by_url = {p["url"]: p for p in [
            {"url": url, "keyword": m["keyword"], "page_type": infer_page_type(url, m.get("intent", "")),
             "current_title": scraped.get(url, {}).get("title", ""),
             "current_h1": scraped.get(url, {}).get("h1", ""),
             "current_meta_desc": scraped.get(url, {}).get("meta_description", "")}
            for url, m in mapping.items() if m.get("lang") == lang
        ]}

        for r in results:
            url = r["url"]
            tags = scraped.get(url, {})
            m = mapping.get(url, {})

            page_data = {
                "url": url,
                "keyword": m.get("keyword", ""),
                "volume": m.get("volume", ""),
                "position": m.get("position", ""),
                "old_title": tags.get("title", ""),
                "new_title": r["new_title"],
                "old_h1": tags.get("h1", ""),
                "new_h1": r["new_h1"],
                "old_meta_desc": tags.get("meta_description", ""),
                "new_meta_desc": r["new_meta_desc"],
            }

            display_rewrite_preview(page_data)
            action = Prompt.ask("[A]ccept / [R]egenerate / [E]dit / [S]kip", default="A")

            if action.upper() == "A":
                export_data[lang].append(page_data)
            elif action.upper() == "R":
                regen_count = 0
                while regen_count < MAX_REGEN_ATTEMPTS:
                    regen_count += 1
                    guidance = Prompt.ask("Guidance (optional, Enter to skip)", default="")
                    page_payload = pages_by_url.get(url, {})
                    new_r = regenerate_single_page(page_payload, brand_name, lang, config, guidance)
                    if new_r:
                        page_data["new_title"] = new_r["new_title"]
                        page_data["new_h1"] = new_r["new_h1"]
                        page_data["new_meta_desc"] = new_r["new_meta_desc"]
                        display_rewrite_preview(page_data)
                        accept = Confirm.ask("Accept this version?")
                        if accept:
                            export_data[lang].append(page_data)
                            break
                    if regen_count >= MAX_REGEN_ATTEMPTS:
                        console.print("[yellow]Max regenerations reached. Edit manually or skip.[/yellow]")
                        manual = Prompt.ask("Edit manually? [Y/n]", default="Y")
                        if manual.upper() == "Y":
                            page_data["new_title"] = Prompt.ask("New title", default=page_data["new_title"])
                            page_data["new_h1"] = Prompt.ask("New H1", default=page_data["new_h1"])
                            page_data["new_meta_desc"] = Prompt.ask("New meta desc", default=page_data["new_meta_desc"])
                            export_data[lang].append(page_data)
            elif action.upper() == "E":
                page_data["new_title"] = Prompt.ask("New title", default=page_data["new_title"])
                page_data["new_h1"] = Prompt.ask("New H1", default=page_data["new_h1"])
                page_data["new_meta_desc"] = Prompt.ask("New meta desc", default=page_data["new_meta_desc"])
                export_data[lang].append(page_data)

    # --- Step 8: Export ---
    server.complete_step("rewriting")
    server.update_progress("review", 0, 0)
    server.complete_step("review")
    today = datetime.date.today().isoformat()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, "..", client_name, "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"{client_name}_optimization_plan_{today}.xlsx")

    console.print(f"\n[bold]Step 8: Exporting to {output_path}...[/bold]")
    server.update_progress("export", 0, 1)
    export_xlsx(export_data, output_path)
    server.update_progress("export", 1, 1)
    server.complete_step("export")
    server.mark_done(grace_seconds=10)
    server.shutdown()
    console.print(f"\n[bold green]✓ Done! Output saved to {output_path}[/bold green]")


if __name__ == "__main__":
    main()
